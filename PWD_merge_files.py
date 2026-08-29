import sys
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters

def safe_key(val1, val2):
    return (str(val1).strip() if val1 is not None else "", str(val2).strip() if val2 is not None else "")

def merge_excel_and_csv(excel_file, csv_file, output_file):
    wb = load_workbook(excel_file, data_only=False)
    ws = wb.active

    excel_data = {}
    excel_headers = [cell.value for cell in ws[1]]
    max_excel_col = ws.max_column

    for row in ws.iter_rows(min_row=2, values_only=False):
        key = safe_key(row[0].value, row[1].value)
        excel_data[key] = row

    with open(csv_file, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        csv_rows = list(reader)
        csv_headers = csv_rows[0]
        csv_data = {}
        for row in csv_rows[1:]:
            key = safe_key(row[0], row[1])
            csv_data[key] = row

    all_keys = sorted(set(excel_data.keys()).union(set(csv_data.keys())))

    new_wb = Workbook()
    new_ws = new_wb.active

    start_csv_col = 19  # Column S
    for col_idx, header in enumerate(excel_headers, start=1):
        new_ws.cell(row=1, column=col_idx, value=header)
    for col_idx, header in enumerate(csv_headers, start=start_csv_col):
        new_ws.cell(row=1, column=col_idx, value=header)
    new_ws.cell(row=1, column=start_csv_col + len(csv_headers), value="Change Type")

    for row_idx, key in enumerate(all_keys, start=2):
        excel_row = excel_data.get(key)
        csv_row = csv_data.get(key)

        # Write Excel data
        if excel_row:
            for col_idx, cell in enumerate(excel_row, start=1):
                new_cell = new_ws.cell(row=row_idx, column=col_idx)
                if cell.data_type == 'f':
                    new_cell.value = f"={cell.value}"
                else:
                    new_cell.value = cell.value
        else:
            for col_idx in range(1, max_excel_col + 1):
                new_ws.cell(row=row_idx, column=col_idx, value=None)

        # Write CSV data
        if csv_row:
            for col_idx, value in enumerate(csv_row, start=start_csv_col):
                new_ws.cell(row=row_idx, column=col_idx, value=value)

        # Determine Change Type
        if excel_row and csv_row:
            excel_password = str(excel_row[2].value).strip() if excel_row[2].value is not None else ""
            csv_password = str(csv_row[2]).strip() if len(csv_row) > 2 else ""
            change_type = "SAME" if excel_password == csv_password else "DIFF"
        elif csv_row and not excel_row:
            change_type = "ADDED"
        elif excel_row and not csv_row:
            change_type = "DELETED"
        else:
            change_type = "UNKNOWN"

        new_ws.cell(row=row_idx, column=start_csv_col + len(csv_headers), value=change_type)

    # Set column widths based on the image
    column_widths = {
        'A': 30,   # url
        'B': 15,   # username
        'C': 15,   # password
        'N': 8,    # Fix
        'O': 10,   # wrong
        'P': 10,   # M_Freq
        'Q': 8,    # pwd
        'R': 8,    # (empty column)
        'S': 30,   # url
        'T': 15,   # username
        'U': 15,   # password
        'AB': 15,  # Change Type
        'AC': 8    # (next column)
    }
    
    for col_letter, width in column_widths.items():
        new_ws.column_dimensions[col_letter].width = width

    # Apply column visibility - show only A, B, C, N, O, P, Q, R, S, T, U, AB
    visible_columns = {1, 2, 3, 14, 15, 16, 17, 18, 19, 20, 21, 28}
    
    max_col = new_ws.max_column
    for col_idx in range(1, max_col + 1):
        if col_idx not in visible_columns:
            col_letter = get_column_letter(col_idx)
            new_ws.column_dimensions[col_letter].hidden = True

    # Apply autofilter to the entire data range
    new_ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{len(all_keys) + 1}"
    
    # Set filter on column AB (28) to hide "SAME" values
    # Column AB is the 28th column, so in the filter it's index 27 (0-based)
    filter_column = FilterColumn(colId=27)
    filter_column.filters = Filters(blank=False)
    # Add the values we want to SHOW
    filter_column.filters.filter = ['ADDED', 'DELETED', 'DIFF', 'UNKNOWN']
    new_ws.auto_filter.filterColumn.append(filter_column)

    new_wb.save(output_file)
    print(f"✅ Merged file saved as: {output_file}")
    print(f"📊 Filter applied to column AB: 'SAME' rows are filtered out (you can toggle this in Excel)")
    print(f"👁️ Visible columns: A, B, C, N, O, P, Q, R, S, T, U, AB")

# === USAGE ===
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python merge_files.py <modified.xlsx> <new.csv> [output.xlsx]")
    else:
        excel_file = sys.argv[1]
        csv_file = sys.argv[2]
        output_file = sys.argv[3] if len(sys.argv) > 3 else "merged_output.xlsx"
        merge_excel_and_csv(excel_file, csv_file, output_file)
